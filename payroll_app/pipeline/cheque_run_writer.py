"""
cheque_run_writer.py — Write reconciled payroll data to PayrollChequeRun_v00.xlsm.

Responsible for:
  1. Writing submitted timesheet hours into columns D–L of 'Time Log'.
  2. Writing final approved hours into columns N–Q of 'Time Log'.
  3. Writing the employee's effective wage rate into column W.
  4. Preserving formula columns X:AC (do NOT overwrite).
  5. Exporting payroll data from 'Current' → V1:AC50 to the Sage 50 CSV path.

Safety rules
------------
  - The workbook is opened with keep_vba=True (VBA must not be removed).
  - Formula columns X:AC are never written to; the writer skips them entirely.
  - The workbook is always saved back to the same path — no temp file gymnastics.
  - The Sage 50 CSV path is non-negotiable: config.SAGE50_PAYROLL_CSV_PATH.
  - All writes are idempotent: re-running for the same period overwrites the
    same rows with the same values.

Time Log row mapping
--------------------
  Row 4 onwards holds one row per employee.
  The writer locates employees by searching column A (employee name) or column B
  (Centerline employee ID).  It does not assume a fixed row per employee.

  Columns D–L  (submitted hours, 9 columns):
    D = REG submitted
    E = OT1 submitted
    F = OT2 submitted
    G = Drive submitted
    H = Sick submitted
    I = Vacation submitted
    J = Holiday submitted
    K = Non-billable submitted
    L = Total submitted (formula — skip)

  Columns N–Q  (approved hours, 4 columns):
    N = REG approved
    O = OT approved
    P = DBL approved
    Q = Travel approved

  Column W = wage rate

Public entry points
-------------------
  write_cheque_run(conn, pay_period_id, *, workbook_path=None, dry_run=False)
      Write all approved reconciliation rows to the Time Log.
      Returns a ChequeRunResult.

  export_sage50_csv(conn, pay_period_id, *, period_end_date, workbook_path=None)
      Read V1:AC50 from the 'Current' sheet and write the Sage 50 CSV.
      Returns the path written to.
"""

import csv
import dataclasses
from pathlib import Path
from typing import Any

import openpyxl

from payroll_app import config
from payroll_app.database import db


# ---------------------------------------------------------------------------
# Workbook constants
# ---------------------------------------------------------------------------

_TIMELOG_SHEET    = "Time Log"
_CURRENT_SHEET    = "Current"

# Column indices (1-based for openpyxl)
_COL_EMP_NAME     = 1    # A
_COL_EMP_ID       = 2    # B  (Centerline employee ID)

# Submitted hours (D=4 through K=11; L=12 is formula — skip)
_COL_D = 4   # REG submitted
_COL_E = 5   # OT1 submitted
_COL_F = 6   # OT2 submitted
_COL_G = 7   # Drive submitted
_COL_H = 8   # Sick submitted
_COL_I = 9   # Vacation submitted
_COL_J = 10  # Holiday submitted
_COL_K = 11  # Non-billable submitted

# Approved hours (N=14 through Q=17)
_COL_N = 14  # REG approved
_COL_O = 15  # OT approved
_COL_P = 16  # DBL approved
_COL_Q = 17  # Travel approved

# Wage rate
_COL_W = 23  # W

# Formula columns — never write to these
_FORMULA_COLS = {24, 25, 26, 27, 28, 29}   # X=24 through AC=29

# First data row in Time Log (header is above row 4)
_TIMELOG_FIRST_ROW = 4
_TIMELOG_LAST_ROW  = 60   # scan up to row 60 for employee name lookups

# Current sheet export range (from config)
# V1:AC50 → columns V=22 through AC=29, rows 1–50
_EXPORT_COL_START = 22   # V
_EXPORT_COL_END   = 29   # AC
_EXPORT_ROW_START = 1
_EXPORT_ROW_END   = 50


# ---------------------------------------------------------------------------
# Result types
# ---------------------------------------------------------------------------

@dataclasses.dataclass
class ChequeRunResult:
    """Result of write_cheque_run()."""
    success:         bool
    pay_period_id:   int
    employees_written: int
    employees_skipped: int   # employee in reconciliation but not found in workbook
    warnings:        list[str]
    errors:          list[str]
    dry_run:         bool


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def _build_employee_row_map(ws) -> dict[str, int]:
    """Scan the Time Log sheet and return a mapping of employee name → row number.

    Looks at column A (display name) and column B (Centerline ID string).
    Both are stored in the map so callers can try either key.
    """
    row_map: dict[str, int] = {}
    for row_num in range(_TIMELOG_FIRST_ROW, _TIMELOG_LAST_ROW + 1):
        name_cell = ws.cell(row=row_num, column=_COL_EMP_NAME).value
        id_cell   = ws.cell(row=row_num, column=_COL_EMP_ID).value

        if name_cell:
            row_map[str(name_cell).strip()] = row_num
        if id_cell:
            row_map[str(id_cell).strip()] = row_num

    return row_map


def _find_employee_row(
    ws,
    display_name: str,
    centerline_id: int | None,
) -> int | None:
    """Return the Time Log row number for an employee, or None if not found."""
    row_map = _build_employee_row_map(ws)

    # Try display name first
    if display_name in row_map:
        return row_map[display_name]

    # Try Centerline ID as string (e.g. "8190" or "E8190")
    if centerline_id is not None:
        for key in (str(centerline_id), f"E{centerline_id}"):
            if key in row_map:
                return row_map[key]

    return None


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_cheque_run(
    conn: Any,
    pay_period_id: int,
    *,
    workbook_path: str | Path | None = None,
    dry_run: bool = False,
) -> ChequeRunResult:
    """Write reconciled payroll data to the Time Log sheet.

    For each employee with a reconciliation row in 'approved' or 'pending' status,
    writes submitted hours (D–K), approved hours (N–Q), and wage rate (W).

    Formula columns (X:AC) are never touched.

    The workbook is opened with keep_vba=True and saved in-place.

    Args:
        conn:           Open database connection.
        pay_period_id:  pay_periods.id — must have reconciliation rows.
        workbook_path:  Path to PayrollChequeRun_v00.xlsm.
                        Defaults to config.PAYROLL_CHEQUE_RUN_WORKBOOK.
        dry_run:        If True, read and compute but do not write to the workbook.

    Returns:
        ChequeRunResult
    """
    warnings: list[str] = []
    errors:   list[str] = []

    wb_path = Path(workbook_path or config.PAYROLL_CHEQUE_RUN_WORKBOOK)

    # Fetch all reconciliation rows for this period (approved + pending)
    rec_rows = db.fetch_all(
        conn,
        """
        SELECT r.*,
               e.display_name, e.centerline_id, e.pdf_id,
               th.reg_hours  AS ts_reg_h,
               th.ot1_hours  AS ts_ot1_h,
               th.ot2_hours  AS ts_ot2_h,
               th.drive_hours AS ts_drv_h,
               th.sick_hours  AS ts_sick_h,
               th.vacation_hours AS ts_vac_h,
               th.holiday_hours  AS ts_hol_h,
               th.nonbillable_hours AS ts_nb_h
        FROM reconciliation r
        JOIN employees e ON e.id = r.employee_id
        LEFT JOIN timesheet_hours th
               ON th.pay_period_id = r.pay_period_id
              AND th.employee_id   = r.employee_id
        WHERE r.pay_period_id = ?
          AND r.status IN ('pending', 'approved')
        ORDER BY e.display_name
        """,
        (pay_period_id,),
    )

    if not rec_rows:
        warnings.append(
            f"No approved/pending reconciliation rows found for pay_period_id={pay_period_id}."
        )
        return ChequeRunResult(
            success=True,
            pay_period_id=pay_period_id,
            employees_written=0,
            employees_skipped=0,
            warnings=warnings,
            errors=errors,
            dry_run=dry_run,
        )

    if dry_run:
        return ChequeRunResult(
            success=True,
            pay_period_id=pay_period_id,
            employees_written=len(rec_rows),
            employees_skipped=0,
            warnings=warnings,
            errors=errors,
            dry_run=True,
        )

    # Workbook existence check — only needed when actually writing
    if not wb_path.exists():
        return ChequeRunResult(
            success=False,
            pay_period_id=pay_period_id,
            employees_written=0,
            employees_skipped=0,
            warnings=warnings,
            errors=[f"Workbook not found: {wb_path}"],
            dry_run=False,
        )

    # Open the workbook (keep_vba=True is required)
    try:
        wb = openpyxl.load_workbook(str(wb_path), keep_vba=True)
    except Exception as exc:
        return ChequeRunResult(
            success=False,
            pay_period_id=pay_period_id,
            employees_written=0,
            employees_skipped=0,
            warnings=warnings,
            errors=[f"Failed to open workbook: {exc}"],
            dry_run=False,
        )

    if _TIMELOG_SHEET not in wb.sheetnames:
        return ChequeRunResult(
            success=False,
            pay_period_id=pay_period_id,
            employees_written=0,
            employees_skipped=0,
            warnings=warnings,
            errors=[f"Sheet {_TIMELOG_SHEET!r} not found in workbook."],
            dry_run=False,
        )

    ws = wb[_TIMELOG_SHEET]
    employees_written  = 0
    employees_skipped  = 0

    for row in rec_rows:
        display_name   = row["display_name"]
        centerline_id  = row["centerline_id"]

        target_row = _find_employee_row(ws, display_name, centerline_id)
        if target_row is None:
            warnings.append(
                f"{display_name}: not found in Time Log — skipped.  "
                "Add employee to the workbook and re-run."
            )
            employees_skipped += 1
            continue

        # Submitted hours (D–K) — from timesheet_hours (biweekly totals)
        ws.cell(row=target_row, column=_COL_D).value = float(row["ts_reg_h"]  or 0)
        ws.cell(row=target_row, column=_COL_E).value = float(row["ts_ot1_h"]  or 0)
        ws.cell(row=target_row, column=_COL_F).value = float(row["ts_ot2_h"]  or 0)
        ws.cell(row=target_row, column=_COL_G).value = float(row["ts_drv_h"]  or 0)
        ws.cell(row=target_row, column=_COL_H).value = float(row["ts_sick_h"] or 0)
        ws.cell(row=target_row, column=_COL_I).value = float(row["ts_vac_h"]  or 0)
        ws.cell(row=target_row, column=_COL_J).value = float(row["ts_hol_h"]  or 0)
        ws.cell(row=target_row, column=_COL_K).value = float(row["ts_nb_h"]   or 0)
        # Column L (total) is a formula — do NOT write

        # Approved hours (N–Q) — from reconciliation final values
        ws.cell(row=target_row, column=_COL_N).value = float(row["final_reg"]   or 0)
        ws.cell(row=target_row, column=_COL_O).value = float(row["final_ot"]    or 0)
        ws.cell(row=target_row, column=_COL_P).value = float(row["final_dbl"]   or 0)
        ws.cell(row=target_row, column=_COL_Q).value = float(row["final_drive"] or 0)

        # Wage rate (W) — look up from employee_rates; leave blank if not found
        rate_row = db.fetch_one(
            conn,
            """
            SELECT base_rate FROM employee_rates
            WHERE employee_id = ?
            ORDER BY effective_date DESC
            LIMIT 1
            """,
            (row["employee_id"],),
        )
        if rate_row and rate_row["base_rate"] is not None:
            ws.cell(row=target_row, column=_COL_W).value = float(rate_row["base_rate"])

        employees_written += 1

    # Save — never strip VBA
    try:
        wb.save(str(wb_path))
    except Exception as exc:
        return ChequeRunResult(
            success=False,
            pay_period_id=pay_period_id,
            employees_written=employees_written,
            employees_skipped=employees_skipped,
            warnings=warnings,
            errors=[f"Failed to save workbook: {exc}"],
            dry_run=False,
        )

    db.log_audit(
        conn,
        action="write_cheque_run",
        entity_type="pay_periods",
        entity_id=pay_period_id,
        new_value=(
            f"employees_written={employees_written}, "
            f"employees_skipped={employees_skipped}, workbook={wb_path.name}"
        ),
    )

    return ChequeRunResult(
        success=True,
        pay_period_id=pay_period_id,
        employees_written=employees_written,
        employees_skipped=employees_skipped,
        warnings=warnings,
        errors=errors,
        dry_run=False,
    )


def export_sage50_csv(
    conn: Any,
    pay_period_id: int,
    *,
    period_end_date: str,
    workbook_path: str | Path | None = None,
) -> Path:
    """Generate the Sage 50 payroll CSV directly from the database and write it.

    Builds one block of 6 rows per approved employee (Regular, Overtime 1,
    Overtime 2, Drive, Holiday, Non-Billable) using final approved hours from
    the reconciliation table and holiday/non-billable hours from timesheet_hours.

    The CSV is written UTF-16 LE with BOM (encoding="utf-16"), which is the
    format Sage 50 requires.  Python's "utf-16" codec writes the BOM
    automatically and selects LE on little-endian systems (Windows).

    Only employees whose reconciliation status is 'approved' or 'exported' are
    included.  Employees are sorted by display_name (or sage50_name alias when
    one is registered).

    The period_end_date must be in YYYYMMDD format (e.g. '20260329').
    The output date column uses M/D/YYYY format without zero-padding.

    The output path is non-negotiable: config.sage50_csv_filename(period_end_date).

    Args:
        conn:             Open database connection.
        pay_period_id:    pay_periods.id
        period_end_date:  YYYYMMDD string for the filename and the Date column.
        workbook_path:    Unused — kept for interface compatibility.

    Returns:
        Path to the CSV file written.

    Raises:
        ValueError: If period_end_date cannot be parsed.
    """
    import datetime as _dt

    # Parse period_end_date (YYYYMMDD) → M/D/YYYY for the Sage 50 Date column
    parsed_end = _dt.date(
        int(period_end_date[0:4]),
        int(period_end_date[4:6]),
        int(period_end_date[6:8]),
    )
    sage50_date = f"{parsed_end.month}/{parsed_end.day}/{parsed_end.year}"

    # Fetch all approved/exported reconciliation rows, joined to employee names
    rec_rows = db.fetch_all(
        conn,
        """
        SELECT r.employee_id, e.display_name,
               r.final_reg, r.final_ot, r.final_dbl, r.final_drive
        FROM reconciliation r
        JOIN employees e ON e.id = r.employee_id
        WHERE r.pay_period_id = ?
          AND r.status IN ('approved', 'exported')
        ORDER BY e.display_name
        """,
        (pay_period_id,),
    )

    # Build the CSV rows: header + 6 rows per employee
    _INCOME_TYPES = [
        "Regular",
        "Overtime 1",
        "Overtime 2",
        "Drive",
        "Holiday",
        "Non-Billable",
    ]

    output_rows: list[list[Any]] = [
        ["Name", "Date", "Income", "Hours", "Pieces", "Amount", "Project", "Comment"]
    ]

    for rec in rec_rows:
        emp_id       = rec["employee_id"]
        display_name = rec["display_name"]

        # Look up sage50_name alias — use it as the Name column if present
        alias_row = db.fetch_one(
            conn,
            """
            SELECT alias_value FROM employee_aliases
            WHERE employee_id = ? AND alias_type = 'sage50_name'
            LIMIT 1
            """,
            (emp_id,),
        )
        sage50_name = alias_row["alias_value"] if alias_row else display_name

        # Fetch holiday and non-billable from timesheet_hours for this period
        ts_row = db.fetch_one(
            conn,
            """
            SELECT holiday_hours, nonbillable_hours
            FROM timesheet_hours
            WHERE pay_period_id = ? AND employee_id = ?
            """,
            (pay_period_id, emp_id),
        )
        holiday_hours     = float(ts_row["holiday_hours"]     or 0) if ts_row else 0.0
        nonbillable_hours = float(ts_row["nonbillable_hours"] or 0) if ts_row else 0.0

        # Hours in the same order as _INCOME_TYPES
        hours_by_type = {
            "Regular":      float(rec["final_reg"]   or 0),
            "Overtime 1":   float(rec["final_ot"]    or 0),
            "Overtime 2":   float(rec["final_dbl"]   or 0),
            "Drive":        float(rec["final_drive"] or 0),
            "Holiday":      holiday_hours,
            "Non-Billable": nonbillable_hours,
        }

        for income_type in _INCOME_TYPES:
            hours = hours_by_type[income_type]
            # Format: drop .0 for whole numbers so "78" not "78.0"
            hours_str = str(int(hours)) if hours == int(hours) else str(hours)
            output_rows.append([
                sage50_name,
                sage50_date,
                income_type,
                hours_str,
                "",   # Pieces
                "",   # Amount
                "",   # Project
                "",   # Comment
            ])

    # Write CSV — UTF-16 LE with BOM (Sage 50 requirement)
    csv_path = config.sage50_csv_filename(period_end_date)
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    with open(str(csv_path), "w", newline="", encoding="utf-16") as fh:
        writer = csv.writer(fh)
        writer.writerows(output_rows)

    db.log_audit(
        conn,
        action="export_sage50_csv",
        entity_type="pay_periods",
        entity_id=pay_period_id,
        new_value=f"csv_path={csv_path}, period_end={period_end_date}",
    )

    return csv_path
