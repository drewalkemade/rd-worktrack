"""
profit_tracker_writer.py — Write weekly approved data to the billing tracker workbook.

Background
----------
The original target was Centerline Profit - 2026.xlsm, but the Phase 1 validation
gate FAILED: openpyxl cannot open the file due to an internal compatibility issue
(Nested.from_tree() error, likely a pivot table or cache incompatibility).

Decision: this writer targets a rebuilt .xlsx file alongside the original .xlsm.
The .xlsm remains the reference; the .xlsx is written by automation.

What it writes
--------------
For each employee in a weekly_approval, one row in the 'RawData' sheet with:
  - Week-ending date
  - Employee name (Centerline display format)
  - Centerline employee ID
  - REG, OT, DBL approved hours
  - Travel hours (Mon–Sat current week)
  - Per-diem count for the week
  - Expense indicator (has other expenses this week?)

The writer appends to the sheet (does not overwrite existing rows).
Re-running for the same weekly_approval replaces only the rows for that approval
(matched by week_ending + employee_id written in fixed columns).

Preserved cells / columns
--------------------------
The .xlsm had formula columns that must not be overwritten.  For the rebuilt
.xlsx, we simply do not write past column R (we stop at column N for approved
travel; formula columns P+ are blank in the rebuilt file).

The protected cell equivalent (AE1 in the original) is written as a named cell
in the rebuilt .xlsx to record the last automation write timestamp.

Public entry points
-------------------
  write_rawdata_week(conn, weekly_approval_id, *, workbook_path=None, dry_run=False)
      Write or update the RawData rows for one weekly_approval.
      Returns a ProfitTrackerResult.

  create_rebuilt_workbook(target_path)
      Create a fresh .xlsx with the RawData sheet and column headers.
      Use this once to initialise the rebuilt workbook.
"""

import dataclasses
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill

from payroll_app import config
from payroll_app.database import db


# ---------------------------------------------------------------------------
# Workbook constants
# ---------------------------------------------------------------------------

_RAWDATA_SHEET  = "RawData"
_LAST_WRITE_CELL = "AE1"   # timestamp of last automation write (mirrors the .xlsm AE1)

# Column layout for RawData (1-based)
_COL_WEEK_ENDING    = 1   # A
_COL_EMP_NAME       = 2   # B
_COL_EMP_ID         = 3   # C  (Centerline ID, e.g. 8190)
_COL_REG            = 4   # D
_COL_OT             = 5   # E
_COL_DBL            = 6   # F
_COL_TRAVEL         = 7   # G  (Mon–Sat current week total)
_COL_PER_DIEM       = 8   # H  (per-diem day count)
_COL_HAS_EXPENSES   = 9   # I  (1 if other expenses present, 0 otherwise)
_COL_WEEKLY_APPR_ID = 10  # J  (internal: weekly_approval_id, for idempotent updates)

# First data row (row 1 = header)
_RAWDATA_FIRST_ROW = 2
_RAWDATA_HEADER = [
    "Week Ending", "Employee", "ID",
    "REG", "OT", "DBL", "Travel",
    "Per Diem Days", "Other Expenses",
    "ApprovalID",   # internal tracking column
]

# Fill colour for "has other expenses" rows — light yellow (mirrors the brown highlight intent)
_EXPENSE_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")


# ---------------------------------------------------------------------------
# Result types
# ---------------------------------------------------------------------------

@dataclasses.dataclass
class ProfitTrackerResult:
    """Result of write_rawdata_week()."""
    success:           bool
    weekly_approval_id: int
    week_ending:       str
    employees_written: int
    employees_skipped: int
    warnings:          list[str]
    errors:            list[str]
    dry_run:           bool


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def _find_or_append_row(ws, weekly_approval_id: int, employee_id: int) -> int:
    """Find an existing row for (weekly_approval_id, employee_id) or return the next empty row.

    Searches column J (ApprovalID) and column C (employee_id via Centerline ID).
    Returns the 1-based row number to write to.
    """
    # The internal approval ID is stored in column J
    for row_num in range(_RAWDATA_FIRST_ROW, ws.max_row + 1):
        cell_appr = ws.cell(row=row_num, column=_COL_WEEKLY_APPR_ID).value
        cell_emp  = ws.cell(row=row_num, column=_COL_EMP_ID).value
        if cell_appr == weekly_approval_id and cell_emp == employee_id:
            return row_num

    # No existing row — use next empty row
    return ws.max_row + 1


def create_rebuilt_workbook(target_path: str | Path) -> Path:
    """Create a fresh rebuilt .xlsx with the RawData sheet and column headers.

    This is a one-time initialisation step.  Run it once to create the file;
    subsequent writes use write_rawdata_week().

    Args:
        target_path: Path to the new .xlsx file to create.

    Returns:
        Path to the created file.
    """
    target_path = Path(target_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _RAWDATA_SHEET

    # Write header row
    for col_idx, header_text in enumerate(_RAWDATA_HEADER, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = Font(bold=True)

    # Record creation timestamp in AE1
    ws[_LAST_WRITE_CELL] = f"Created {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    # Freeze the top row
    ws.freeze_panes = "A2"

    target_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(target_path))
    return target_path


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_rawdata_week(
    conn: Any,
    weekly_approval_id: int,
    *,
    workbook_path: str | Path | None = None,
    dry_run: bool = False,
) -> ProfitTrackerResult:
    """Write or update RawData rows for one weekly_approval.

    For each employee in the weekly_approval with customer_hours:
      - Locates their existing row in RawData (by weekly_approval_id + employee_id)
        or appends a new row.
      - Writes week_ending, employee name/ID, approved hours, travel, and expense flags.
      - Applies a light-yellow fill if the employee has non-per-diem expenses this week
        (mirrors the manual brown-highlight intent from the original workflow).

    Updates AE1 with the write timestamp.

    Args:
        conn:                Open database connection.
        weekly_approval_id:  weekly_approvals.id to write.
        workbook_path:       Path to the rebuilt .xlsx.
                             Defaults to a 'Centerline_Profit_Rebuilt.xlsx' alongside
                             the original .xlsm in the project root.
        dry_run:             If True, compute but do not write.

    Returns:
        ProfitTrackerResult
    """
    warnings: list[str] = []
    errors:   list[str] = []

    # Default workbook path: Centerline_Profit_Rebuilt.xlsx in project root
    if workbook_path is None:
        workbook_path = config.PROJECT_ROOT / "Centerline_Profit_Rebuilt.xlsx"
    wb_path = Path(workbook_path)

    wa = db.fetch_one(
        conn,
        "SELECT * FROM weekly_approvals WHERE id = ?",
        (weekly_approval_id,),
    )
    if not wa:
        return ProfitTrackerResult(
            success=False,
            weekly_approval_id=weekly_approval_id,
            week_ending="",
            employees_written=0,
            employees_skipped=0,
            warnings=warnings,
            errors=[f"weekly_approval_id={weekly_approval_id} not found."],
            dry_run=dry_run,
        )

    week_ending   = wa["week_ending"]
    pay_period_id = wa["pay_period_id"]

    # Fetch all employees for this approval
    rows = db.fetch_all(
        conn,
        """
        SELECT ch.employee_id, ch.reg_hours, ch.ot_hours, ch.dbl_hours,
               e.display_name, e.centerline_id,
               th.current_week_total AS travel_hours
        FROM customer_hours ch
        JOIN employees e ON e.id = ch.employee_id
        LEFT JOIN travel_hours th
               ON th.weekly_approval_id = ch.weekly_approval_id
              AND th.employee_id = ch.employee_id
        WHERE ch.weekly_approval_id = ?
        ORDER BY e.display_name
        """,
        (weekly_approval_id,),
    )

    if not rows:
        warnings.append(f"No customer_hours found for weekly_approval_id={weekly_approval_id}.")
        return ProfitTrackerResult(
            success=True,
            weekly_approval_id=weekly_approval_id,
            week_ending=week_ending,
            employees_written=0,
            employees_skipped=0,
            warnings=warnings,
            errors=errors,
            dry_run=dry_run,
        )

    if dry_run:
        return ProfitTrackerResult(
            success=True,
            weekly_approval_id=weekly_approval_id,
            week_ending=week_ending,
            employees_written=len(rows),
            employees_skipped=0,
            warnings=warnings,
            errors=errors,
            dry_run=True,
        )

    # Create workbook if it does not yet exist
    if not wb_path.exists():
        create_rebuilt_workbook(wb_path)

    try:
        wb = openpyxl.load_workbook(str(wb_path))
    except Exception as exc:
        return ProfitTrackerResult(
            success=False,
            weekly_approval_id=weekly_approval_id,
            week_ending=week_ending,
            employees_written=0,
            employees_skipped=0,
            warnings=warnings,
            errors=[f"Failed to open workbook {wb_path}: {exc}"],
            dry_run=False,
        )

    if _RAWDATA_SHEET not in wb.sheetnames:
        return ProfitTrackerResult(
            success=False,
            weekly_approval_id=weekly_approval_id,
            week_ending=week_ending,
            employees_written=0,
            employees_skipped=0,
            warnings=warnings,
            errors=[f"Sheet {_RAWDATA_SHEET!r} not found in {wb_path.name}."],
            dry_run=False,
        )

    ws = wb[_RAWDATA_SHEET]
    week_end_date = date.fromisoformat(week_ending)
    employees_written = 0

    for row in rows:
        employee_id   = row["employee_id"]
        display_name  = row["display_name"]
        centerline_id = row["centerline_id"]

        # Expense flags for this employee/week
        week_start = str(week_end_date - __import__('datetime').timedelta(days=6))
        per_diem_count = 0.0
        has_other_expenses = 0

        exp_rows = db.fetch_all(
            conn,
            """
            SELECT category, amount
            FROM expense_items
            WHERE pay_period_id = ?
              AND employee_id = ?
              AND (work_date IS NULL OR (work_date BETWEEN ? AND ?))
            """,
            (pay_period_id, employee_id, week_start, week_ending),
        )
        for exp in exp_rows:
            if exp["category"] in ("per_diem_travel", "per_diem_full"):
                per_diem_count += 1.0
            elif exp["amount"] and float(exp["amount"]) > 0:
                has_other_expenses = 1

        # Find or append row
        target_row = _find_or_append_row(ws, weekly_approval_id, centerline_id or employee_id)

        ws.cell(row=target_row, column=_COL_WEEK_ENDING).value    = week_end_date
        ws.cell(row=target_row, column=_COL_EMP_NAME).value       = display_name
        ws.cell(row=target_row, column=_COL_EMP_ID).value         = centerline_id
        ws.cell(row=target_row, column=_COL_REG).value            = float(row["reg_hours"]  or 0)
        ws.cell(row=target_row, column=_COL_OT).value             = float(row["ot_hours"]   or 0)
        ws.cell(row=target_row, column=_COL_DBL).value            = float(row["dbl_hours"]  or 0)
        ws.cell(row=target_row, column=_COL_TRAVEL).value         = float(row["travel_hours"] or 0)
        ws.cell(row=target_row, column=_COL_PER_DIEM).value       = per_diem_count
        ws.cell(row=target_row, column=_COL_HAS_EXPENSES).value   = has_other_expenses
        ws.cell(row=target_row, column=_COL_WEEKLY_APPR_ID).value = weekly_approval_id

        # Light-yellow fill when other expenses exist (mirrors manual brown highlight)
        if has_other_expenses:
            for col_num in range(1, 11):
                ws.cell(row=target_row, column=col_num).fill = _EXPENSE_FILL

        employees_written += 1

    # Update AE1 with write timestamp
    ws[_LAST_WRITE_CELL] = f"Last written {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    try:
        wb.save(str(wb_path))
    except Exception as exc:
        return ProfitTrackerResult(
            success=False,
            weekly_approval_id=weekly_approval_id,
            week_ending=week_ending,
            employees_written=employees_written,
            employees_skipped=0,
            warnings=warnings,
            errors=[f"Failed to save workbook: {exc}"],
            dry_run=False,
        )

    db.log_audit(
        conn,
        action="write_rawdata_week",
        entity_type="weekly_approvals",
        entity_id=weekly_approval_id,
        new_value=(
            f"week_ending={week_ending}, employees_written={employees_written}, "
            f"workbook={wb_path.name}"
        ),
    )

    return ProfitTrackerResult(
        success=True,
        weekly_approval_id=weekly_approval_id,
        week_ending=week_ending,
        employees_written=employees_written,
        employees_skipped=0,
        warnings=warnings,
        errors=errors,
        dry_run=False,
    )
