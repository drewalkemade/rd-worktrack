"""
test_workbook_writers.py — Tests for the workbook writers and receipt ingest.

These tests cover:
  - profit_tracker_writer: create_rebuilt_workbook, write_rawdata_week (dry_run)
  - cheque_run_writer: write_cheque_run (dry_run + workbook-not-found error path)
  - receipt_ingest: suggest_normalized_name, ingest_receipt (with a stub receipt file)
"""

import sqlite3
from datetime import date
from pathlib import Path

import pytest

from payroll_app import config
from payroll_app.database import db, employee_manager
from payroll_app.pipeline.importer import import_payroll_pdf, import_timesheet
from payroll_app.pipeline.profit_tracker_writer import (
    create_rebuilt_workbook,
    write_rawdata_week,
)
from payroll_app.pipeline.cheque_run_writer import write_cheque_run
from payroll_app.extractors.receipt_ingest import (
    suggest_normalized_name,
    ingest_receipt,
    get_receipts_for_period,
)

FIXTURES = Path(__file__).parent / "fixtures"

PAYROLL_PDF = FIXTURES / "R&D_260329-xxxxx.pdf"
TRIF_TS     = FIXTURES / "DTrifTS_17_20260329.xlsx"

WEEK_ENDING = date(2026, 3, 29)


@pytest.fixture()
def conn(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "SOURCE_FILES_DIR", tmp_path)
    monkeypatch.setattr(config, "PAYROLL_PDF_DIR",  tmp_path / "payroll_pdfs")
    monkeypatch.setattr(config, "TRAVEL_PDF_DIR",   tmp_path / "travel_pdfs")
    monkeypatch.setattr(config, "TIMESHEET_DIR",    tmp_path / "timesheets")
    monkeypatch.setattr(config, "RECEIPT_DIR",      tmp_path / "receipts")

    c = sqlite3.connect(":memory:")
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA foreign_keys = ON")
    db.initialize_database(c)
    employee_manager.seed_employees(c)
    yield c
    c.close()


# ---------------------------------------------------------------------------
# profit_tracker_writer tests
# ---------------------------------------------------------------------------

class TestProfitTrackerWriter:

    def test_create_rebuilt_workbook(self, tmp_path):
        """create_rebuilt_workbook() creates a valid .xlsx with RawData sheet."""
        import openpyxl
        target = tmp_path / "test_profit.xlsx"
        path = create_rebuilt_workbook(target)
        assert path.exists()

        wb = openpyxl.load_workbook(str(path))
        assert "RawData" in wb.sheetnames
        ws = wb["RawData"]
        # Header row should have column names
        assert ws.cell(row=1, column=1).value == "Week Ending"

    def test_write_rawdata_week_dry_run(self, conn, tmp_path):
        """dry_run=True returns a successful result without writing files."""
        pr = import_payroll_pdf(conn, PAYROLL_PDF, WEEK_ENDING)
        wa_id = pr.weekly_approval_id

        result = write_rawdata_week(
            conn, wa_id,
            workbook_path=tmp_path / "profit.xlsx",
            dry_run=True,
        )
        assert result.success is True
        assert result.dry_run is True
        assert result.employees_written == 6   # 6 known employees
        # File should NOT have been created in dry_run mode
        assert not (tmp_path / "profit.xlsx").exists()

    def test_write_rawdata_week_creates_workbook(self, conn, tmp_path):
        """If the target .xlsx does not exist, write_rawdata_week creates it."""
        import openpyxl
        pr = import_payroll_pdf(conn, PAYROLL_PDF, WEEK_ENDING)
        wa_id = pr.weekly_approval_id
        target = tmp_path / "profit_new.xlsx"

        result = write_rawdata_week(conn, wa_id, workbook_path=target)
        assert result.success is True
        assert target.exists()

        wb = openpyxl.load_workbook(str(target))
        ws = wb["RawData"]
        # At least one data row should have been written
        data_rows = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value]
        assert len(data_rows) >= 1

    def test_write_rawdata_week_idempotent(self, conn, tmp_path):
        """Re-running for the same approval updates existing rows rather than appending."""
        import openpyxl
        pr = import_payroll_pdf(conn, PAYROLL_PDF, WEEK_ENDING)
        wa_id = pr.weekly_approval_id
        target = tmp_path / "profit_idem.xlsx"

        write_rawdata_week(conn, wa_id, workbook_path=target)
        write_rawdata_week(conn, wa_id, workbook_path=target)

        wb = openpyxl.load_workbook(str(target))
        ws = wb["RawData"]
        # Count non-empty rows after header
        data_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value]
        # Should have exactly 6 rows (one per employee), not 12
        assert len(data_rows) == 6

    def test_nonexistent_approval_returns_error(self, conn, tmp_path):
        result = write_rawdata_week(conn, weekly_approval_id=9999, workbook_path=tmp_path / "x.xlsx")
        assert result.success is False
        assert result.errors

    def test_audit_log_created(self, conn, tmp_path):
        pr = import_payroll_pdf(conn, PAYROLL_PDF, WEEK_ENDING)
        wa_id = pr.weekly_approval_id
        write_rawdata_week(conn, wa_id, workbook_path=tmp_path / "profit_audit.xlsx")
        row = db.fetch_one(conn, "SELECT * FROM audit_log WHERE action = 'write_rawdata_week'")
        assert row is not None


# ---------------------------------------------------------------------------
# cheque_run_writer tests
# ---------------------------------------------------------------------------

class TestChequeRunWriter:

    def test_workbook_not_found_returns_error(self, conn, tmp_path):
        import_payroll_pdf(conn, PAYROLL_PDF, WEEK_ENDING)
        import_timesheet(conn, TRIF_TS)
        pid = db.fetch_one(conn, "SELECT id FROM pay_periods LIMIT 1")["id"]

        # Insert a reconciliation row so the writer proceeds past the "no rows" check
        conn.execute(
            """
            INSERT INTO reconciliation (pay_period_id, employee_id, ts_reg, ts_ot, ts_dbl, ts_drive,
              cust_reg, cust_ot, cust_dbl, cust_drive, final_reg, final_ot, final_dbl, final_drive, status)
            SELECT ?, id, 0,0,0,0,0,0,0,0,0,0,0,0,'pending'
            FROM employees WHERE pdf_name = 'TRIF, DANIEL'
            """,
            (pid,),
        )

        result = write_cheque_run(conn, pid, workbook_path=tmp_path / "nonexistent.xlsm")
        assert result.success is False
        assert any("not found" in e for e in result.errors)

    def test_dry_run_returns_success(self, conn, tmp_path, monkeypatch):
        """dry_run=True with a non-existent workbook still returns success=True
        because dry_run does not open the file."""
        import_payroll_pdf(conn, PAYROLL_PDF, WEEK_ENDING)
        import_timesheet(conn, TRIF_TS)
        pid = db.fetch_one(conn, "SELECT id FROM pay_periods LIMIT 1")["id"]

        # Insert minimal reconciliation rows so write_cheque_run has data
        conn.execute(
            """
            INSERT INTO reconciliation (pay_period_id, employee_id, ts_reg, ts_ot, ts_dbl, ts_drive,
              cust_reg, cust_ot, cust_dbl, cust_drive, final_reg, final_ot, final_dbl, final_drive, status)
            SELECT ?, id, 0,0,0,0,0,0,0,0,0,0,0,0,'pending'
            FROM employees WHERE pdf_name = 'TRIF, DANIEL'
            """,
            (pid,),
        )

        result = write_cheque_run(
            conn, pid,
            workbook_path=tmp_path / "nonexistent.xlsm",
            dry_run=True,
        )
        assert result.success is True
        assert result.dry_run is True


# ---------------------------------------------------------------------------
# receipt_ingest tests
# ---------------------------------------------------------------------------

class TestSuggestNormalizedName:

    def test_basic_name_derivation(self):
        name = suggest_normalized_name(
            display_name="Daniel Trif",
            category="lodging",
            work_date="2026-03-24",
            original_ext=".jpg",
        )
        assert name == "DTRIF_lodging_2026-03-24.jpg"

    def test_undated_when_no_date(self):
        name = suggest_normalized_name(
            display_name="Daniel Trif",
            category="per_diem_travel",
            work_date=None,
            original_ext=".pdf",
        )
        assert "undated" in name

    def test_category_slug_replaces_spaces(self):
        name = suggest_normalized_name(
            display_name="Florin Moldovan",
            category="car_rental",
            work_date="2026-03-18",
            original_ext=".png",
        )
        assert "car_rental" in name

    def test_expense_code_override(self):
        name = suggest_normalized_name(
            display_name="Daniel Trif",
            category="tolls",
            work_date="2026-03-20",
            original_ext=".jpg",
            expense_code="DTRIF",
        )
        assert name.startswith("DTRIF_")

    def test_extension_lowercased(self):
        name = suggest_normalized_name(
            display_name="Daniel Trif",
            category="lodging",
            work_date="2026-03-24",
            original_ext=".JPG",
        )
        assert name.endswith(".jpg")


class TestIngestReceipt:

    def _make_stub_receipt(self, tmp_path: Path, filename: str = "receipt.jpg") -> Path:
        """Write a small stub file to act as a receipt."""
        receipt = tmp_path / filename
        receipt.write_bytes(b"stub receipt content")
        return receipt

    def test_ingest_success(self, conn, tmp_path):
        """Ingest a receipt against a per-diem expense item (no receipt required, but attachable)."""
        import_timesheet(conn, TRIF_TS)

        # Find a per-diem expense item for Trif
        emp_id = db.fetch_one(conn, "SELECT id FROM employees WHERE pdf_name = 'TRIF, DANIEL'")["id"]
        pid = db.fetch_one(conn, "SELECT id FROM pay_periods LIMIT 1")["id"]
        item = db.fetch_one(
            conn,
            "SELECT id FROM expense_items WHERE employee_id = ? AND pay_period_id = ?",
            (emp_id, pid),
        )
        if not item:
            pytest.skip("No expense items in Trif fixture")

        stub = self._make_stub_receipt(tmp_path)
        result = ingest_receipt(conn, stub, item["id"])

        assert result.success is True
        assert result.source_file_id is not None
        assert result.normalized_name is not None

    def test_ingest_links_receipt_to_expense(self, conn, tmp_path):
        import_timesheet(conn, TRIF_TS)
        emp_id = db.fetch_one(conn, "SELECT id FROM employees WHERE pdf_name = 'TRIF, DANIEL'")["id"]
        pid = db.fetch_one(conn, "SELECT id FROM pay_periods LIMIT 1")["id"]
        item = db.fetch_one(
            conn,
            "SELECT id FROM expense_items WHERE employee_id = ? AND pay_period_id = ?",
            (emp_id, pid),
        )
        if not item:
            pytest.skip("No expense items in Trif fixture")

        stub = self._make_stub_receipt(tmp_path)
        ingest_receipt(conn, stub, item["id"])

        receipt_row = db.fetch_one(
            conn,
            "SELECT * FROM expense_receipts WHERE expense_item_id = ?",
            (item["id"],),
        )
        assert receipt_row is not None
        assert receipt_row["stored_path"] is not None

    def test_ingest_receipt_not_found(self, conn, tmp_path):
        result = ingest_receipt(conn, tmp_path / "nonexistent.jpg", expense_item_id=1)
        assert result.success is False

    def test_ingest_invalid_expense_item(self, conn, tmp_path):
        stub = self._make_stub_receipt(tmp_path)
        result = ingest_receipt(conn, stub, expense_item_id=9999)
        assert result.success is False
        assert any("not found" in e for e in result.errors)

    def test_get_receipts_for_period(self, conn, tmp_path):
        import_timesheet(conn, TRIF_TS)
        emp_id = db.fetch_one(conn, "SELECT id FROM employees WHERE pdf_name = 'TRIF, DANIEL'")["id"]
        pid = db.fetch_one(conn, "SELECT id FROM pay_periods LIMIT 1")["id"]
        item = db.fetch_one(
            conn,
            "SELECT id FROM expense_items WHERE employee_id = ? AND pay_period_id = ?",
            (emp_id, pid),
        )
        if not item:
            pytest.skip("No expense items in Trif fixture")

        stub = self._make_stub_receipt(tmp_path)
        ingest_receipt(conn, stub, item["id"])

        receipts = get_receipts_for_period(conn, pid)
        assert len(receipts) == 1
        assert receipts[0].expense_item_id == item["id"]

    def test_audit_log_created_on_ingest(self, conn, tmp_path):
        import_timesheet(conn, TRIF_TS)
        emp_id = db.fetch_one(conn, "SELECT id FROM employees WHERE pdf_name = 'TRIF, DANIEL'")["id"]
        pid = db.fetch_one(conn, "SELECT id FROM pay_periods LIMIT 1")["id"]
        item = db.fetch_one(
            conn,
            "SELECT id FROM expense_items WHERE employee_id = ? AND pay_period_id = ?",
            (emp_id, pid),
        )
        if not item:
            pytest.skip("No expense items in Trif fixture")

        stub = self._make_stub_receipt(tmp_path)
        ingest_receipt(conn, stub, item["id"])

        row = db.fetch_one(conn, "SELECT * FROM audit_log WHERE action = 'ingest_receipt'")
        assert row is not None
