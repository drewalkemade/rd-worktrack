"""
timesheet_extractor_v2.py — Daily-level timesheet and expense extractor.

Replaces the legacy extractor that read only biweekly totals from row 23.

This version extracts:
  1. Daily labor hours from rows 9–22 of 'Biweekly Time Sheet' (14 rows for 14 days).
     The second Sunday (end of period) is explicitly included — do not assume rows
     end at row 21.
  2. Biweekly totals from row 23 (for cross-checking against the sum of daily rows).
  3. Expense line items from 'Biweekly Expense Report CAD' rows 10–23.
  4. Expense line items from 'Biweekly Expense Report USD' rows 10–23.

Workbook structure (from EmpTS - 20260329.xlsx):
  Biweekly Time Sheet:
    H3  = employee name
    H5  = period end date
    B9:B22  = day names (Monday, Tuesday, …, Sunday)
    C9:C22  = dates (datetime objects)
    D9:D22  = Regular Hours
    E9:E22  = Overtime 1 Hours
    F9:F22  = Overtime 2 Hours
    G9:G22  = Drive Hours
    H9:H22  = Sick/PEL
    I9:I22  = Vacation
    J9:J22  = Holiday
    K9:K22  = Non-Billable
    L9:L22  = Total (formula — cross-check only)
    Row 23  = Totals row (C23="Total hours", D23:K23 = sums)

  Biweekly Expense Report CAD / USD:
    K3  = "Employee:"
    L3  = employee name
    K4  = "Pay period start date:"
    L4  = period start date
    K5  = "Pay period end date:"
    L5  = period end date
    Row 9 = column headers (D=Per Diem - Travel, E=Per Diem - Full, …, N=Mileage)
    Rows 10–23 = daily expense rows (B=Day, C=Date, D:N=amounts)
    Row 24 = Totals row (C24="Total", D24:N24 = sums)

Returns:
  {
    "employee_name":     str,
    "period_start":      date,
    "period_end":        date,
    "source_file":       str,    # original path as provided
    "daily_hours": [
      {
        "row_number":      int,   # spreadsheet row (9–22)
        "day_name":        str,   # "Monday", "Tuesday", etc.
        "work_date":       date,
        "reg_hours":       float,
        "ot1_hours":       float,
        "ot2_hours":       float,
        "drive_hours":     float,
        "sick_hours":      float,
        "vacation_hours":  float,
        "holiday_hours":   float,
        "nonbillable_hours": float,
        "row_total":       float,  # from column L (formula value)
        "computed_total":  float,  # sum of individual hour columns
      }
    ],
    "totals": {
      "reg_hours":         float,
      "ot1_hours":         float,
      "ot2_hours":         float,
      "drive_hours":       float,
      "sick_hours":        float,
      "vacation_hours":    float,
      "holiday_hours":     float,
      "nonbillable_hours": float,
      "sum_total":         float,   # from column L row 23
      "computed_total":    float,   # derived from daily rows
    },
    "totals_match":  bool,   # True if computed_total == row-23 totals
    "warnings":      list[str],
    "expenses_cad": [
      {
        "row_number":   int,
        "day_name":     str,
        "work_date":    date | None,
        "category":     str,   # e.g. "per_diem_travel"
        "amount":       float,
        "currency":     "CAD",
        "requires_receipt": bool,
      }
    ],
    "expenses_usd": [ ... same shape as expenses_cad, currency="USD" ... ],
    "expense_totals_cad": { category: float, ... },
    "expense_totals_usd": { category: float, ... },
  }
"""

from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import openpyxl.utils as xl_utils

from payroll_app import config


# ---------------------------------------------------------------------------
# Timesheet sheet constants
# ---------------------------------------------------------------------------

_TS_SHEET       = "Biweekly Time Sheet"
_TS_NAME_CELL   = "H3"
_TS_END_DATE    = "H5"
_TS_HEADER_ROW  = 8
_TS_FIRST_DAY_ROW = 9
_TS_LAST_DAY_ROW  = 22
_TS_TOTALS_ROW    = 23

# Column letters for the Biweekly Time Sheet
_TS_COL_DAY       = "B"
_TS_COL_DATE      = "C"
_TS_COL_REG       = "D"
_TS_COL_OT1       = "E"
_TS_COL_OT2       = "F"
_TS_COL_DRIVE     = "G"
_TS_COL_SICK      = "H"
_TS_COL_VACATION  = "I"
_TS_COL_HOLIDAY   = "J"
_TS_COL_NONBILL   = "K"
_TS_COL_TOTAL     = "L"

_TS_HOUR_COLS = [
    ("reg_hours",        _TS_COL_REG),
    ("ot1_hours",        _TS_COL_OT1),
    ("ot2_hours",        _TS_COL_OT2),
    ("drive_hours",      _TS_COL_DRIVE),
    ("sick_hours",       _TS_COL_SICK),
    ("vacation_hours",   _TS_COL_VACATION),
    ("holiday_hours",    _TS_COL_HOLIDAY),
    ("nonbillable_hours",_TS_COL_NONBILL),
]

# ---------------------------------------------------------------------------
# Expense sheet constants
# ---------------------------------------------------------------------------

_EXP_SHEET_CAD  = "Biweekly Expense Report CAD"
_EXP_SHEET_USD  = "Biweekly Expense Report USD"

_EXP_EMP_CELL   = "L3"    # employee name
_EXP_START_CELL = "L4"    # period start date
_EXP_END_CELL   = "L5"    # period end date

_EXP_HEADER_ROW = 9
_EXP_FIRST_ROW  = 10
_EXP_LAST_ROW   = 23
_EXP_TOTALS_ROW = 24

_EXP_COL_DAY  = "B"
_EXP_COL_DATE = "C"

# Category labels as they appear in the header row, mapped to our canonical names
_EXP_CATEGORY_MAP = {
    "Per Diem - Travel":       "per_diem_travel",
    "Per Diem - Full":         "per_diem_full",
    "Lodging":                 "lodging",
    "Lodging - Per Receipt":   "lodging_per_receipt",
    "Luggage":                 "luggage",
    "Car Rental":              "car_rental",
    "Tolls":                   "tolls",
    "Vehicle Maintenance":     "vehicle_maintenance",
    "Parking Fees":            "parking_fees",
    "Taxi":                    "taxi",
    "Mileage":                 "mileage",
}

# Columns D through N in the expense sheets
_EXP_DATA_COL_START = 4   # column D = index 4
_EXP_DATA_COL_END   = 14  # column N = index 14


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _safe_float(value: Any) -> float:
    """Convert a cell value to float, defaulting to 0.0."""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def _to_date(value: Any) -> date | None:
    """Convert an openpyxl date/datetime cell value to a Python date, or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def _derive_period_start(period_end: date) -> date:
    """Derive the 14-day period start from the period end date.

    The period always ends on a Sunday (the last of 14 days).
    The period starts on the Monday 13 days earlier.
    """
    from datetime import timedelta
    return period_end - timedelta(days=13)


# ---------------------------------------------------------------------------
# Expense sheet parsing
# ---------------------------------------------------------------------------

def _read_expense_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    currency: str,
    warnings: list[str],
) -> tuple[list[dict], dict[str, float]]:
    """Extract expense rows from one expense sheet (CAD or USD).

    Returns:
        (expense_rows, totals_by_category)
        expense_rows: list of dicts per non-zero expense line item
        totals_by_category: { canonical_category: total_amount }
    """
    if sheet_name not in wb.sheetnames:
        warnings.append(f"Sheet {sheet_name!r} not found in workbook.")
        return [], {}

    ws = wb[sheet_name]
    expense_rows: list[dict] = []
    totals: dict[str, float] = {}

    # Build column → canonical category map from the header row
    col_to_category: dict[int, str] = {}
    for col_idx in range(_EXP_DATA_COL_START, _EXP_DATA_COL_END + 1):
        header_value = ws.cell(row=_EXP_HEADER_ROW, column=col_idx).value
        if header_value is not None:
            label = str(header_value).strip()
            canonical = _EXP_CATEGORY_MAP.get(label)
            if canonical:
                col_to_category[col_idx] = canonical
            else:
                warnings.append(
                    f"{sheet_name}: Unknown expense category header {label!r} "
                    f"in column {xl_utils.get_column_letter(col_idx)}."
                )

    # Read daily rows
    for row_num in range(_EXP_FIRST_ROW, _EXP_LAST_ROW + 1):
        day_name = ws[f"{_EXP_COL_DAY}{row_num}"].value
        date_val = _to_date(ws[f"{_EXP_COL_DATE}{row_num}"].value)

        for col_idx, category in col_to_category.items():
            amount = _safe_float(ws.cell(row=row_num, column=col_idx).value)
            if amount == 0.0:
                continue

            requires_receipt = category not in config.PER_DIEM_CATEGORIES

            expense_rows.append({
                "row_number":        row_num,
                "day_name":          str(day_name).strip() if day_name else "",
                "work_date":         date_val,
                "category":          category,
                "amount":            amount,
                "currency":          currency,
                "requires_receipt":  requires_receipt,
            })

    # Read totals row
    for col_idx, category in col_to_category.items():
        total_val = _safe_float(ws.cell(row=_EXP_TOTALS_ROW, column=col_idx).value)
        totals[category] = total_val

    # Cross-check: sum of daily rows should match totals row
    computed: dict[str, float] = {}
    for row in expense_rows:
        cat = row["category"]
        computed[cat] = computed.get(cat, 0.0) + row["amount"]

    for category, total in totals.items():
        computed_total = computed.get(category, 0.0)
        if abs(computed_total - total) > 0.01:
            warnings.append(
                f"{sheet_name}: Totals row mismatch for {category!r}: "
                f"row 24 says {total:.2f}, sum of daily rows is {computed_total:.2f}."
            )

    return expense_rows, totals


# ---------------------------------------------------------------------------
# Main extractor
# ---------------------------------------------------------------------------

def extract_timesheet(file_path: str | Path) -> dict[str, Any]:
    """Extract all data from an employee timesheet workbook.

    Args:
        file_path: Path to the .xlsx timesheet file.

    Returns:
        A dict matching the shape described in the module docstring.

    Raises:
        FileNotFoundError: If file_path does not exist.
        ValueError: If the required sheet or essential cells are missing.
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Timesheet not found: {file_path}")

    warnings: list[str] = []

    # Load with data_only=True to read formula results, not formula strings
    wb = openpyxl.load_workbook(str(file_path), data_only=True)

    if _TS_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Required sheet {_TS_SHEET!r} not found in {file_path.name}. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[_TS_SHEET]

    # ---- Employee name ----
    employee_name = ws[_TS_NAME_CELL].value
    if not employee_name:
        raise ValueError(f"No employee name in cell {_TS_NAME_CELL} of {file_path.name}.")
    employee_name = str(employee_name).strip()

    # ---- Period end date ----
    period_end = _to_date(ws[_TS_END_DATE].value)
    if period_end is None:
        raise ValueError(f"No valid end date in cell {_TS_END_DATE} of {file_path.name}.")

    period_start = _derive_period_start(period_end)

    # ---- Daily hour rows ----
    daily_hours: list[dict] = []

    for row_num in range(_TS_FIRST_DAY_ROW, _TS_LAST_DAY_ROW + 1):
        day_name = ws[f"{_TS_COL_DAY}{row_num}"].value
        work_date = _to_date(ws[f"{_TS_COL_DATE}{row_num}"].value)

        if day_name is None and work_date is None:
            warnings.append(f"Row {row_num}: Both day name and date are empty — skipping.")
            continue

        hour_values: dict[str, float] = {}
        for field_name, col_letter in _TS_HOUR_COLS:
            hour_values[field_name] = _safe_float(ws[f"{col_letter}{row_num}"].value)

        row_total   = _safe_float(ws[f"{_TS_COL_TOTAL}{row_num}"].value)
        comp_total  = sum(hour_values.values())

        if abs(comp_total - row_total) > 0.01 and row_total > 0:
            warnings.append(
                f"Row {row_num} ({day_name}): Row total ({row_total}) does not match "
                f"sum of individual columns ({comp_total:.2f})."
            )

        daily_hours.append({
            "row_number":        row_num,
            "day_name":          str(day_name).strip() if day_name else "",
            "work_date":         work_date,
            "reg_hours":         hour_values["reg_hours"],
            "ot1_hours":         hour_values["ot1_hours"],
            "ot2_hours":         hour_values["ot2_hours"],
            "drive_hours":       hour_values["drive_hours"],
            "sick_hours":        hour_values["sick_hours"],
            "vacation_hours":    hour_values["vacation_hours"],
            "holiday_hours":     hour_values["holiday_hours"],
            "nonbillable_hours": hour_values["nonbillable_hours"],
            "row_total":         row_total,
            "computed_total":    comp_total,
        })

    if len(daily_hours) != 14:
        warnings.append(
            f"Expected 14 daily rows (rows {_TS_FIRST_DAY_ROW}–{_TS_LAST_DAY_ROW}), "
            f"found {len(daily_hours)}."
        )

    # ---- Biweekly totals row ----
    totals: dict[str, float] = {}
    for field_name, col_letter in _TS_HOUR_COLS:
        totals[field_name] = _safe_float(ws[f"{col_letter}{_TS_TOTALS_ROW}"].value)
    totals["sum_total"] = _safe_float(ws[f"{_TS_COL_TOTAL}{_TS_TOTALS_ROW}"].value)

    # Cross-check: sum daily rows against totals row
    totals["computed_total"] = sum(
        sum(row[f] for f, _ in _TS_HOUR_COLS) for row in daily_hours
    )

    totals_match = True
    for field_name, _ in _TS_HOUR_COLS:
        computed = sum(row[field_name] for row in daily_hours)
        if abs(computed - totals[field_name]) > 0.01:
            totals_match = False
            warnings.append(
                f"Totals row mismatch for {field_name}: "
                f"row {_TS_TOTALS_ROW} says {totals[field_name]:.2f}, "
                f"sum of daily rows is {computed:.2f}."
            )

    # ---- Expense sheets ----
    expenses_cad, totals_cad = _read_expense_sheet(wb, _EXP_SHEET_CAD, "CAD", warnings)
    expenses_usd, totals_usd = _read_expense_sheet(wb, _EXP_SHEET_USD, "USD", warnings)

    return {
        "employee_name":      employee_name,
        "period_start":       period_start,
        "period_end":         period_end,
        "source_file":        str(file_path),
        "daily_hours":        daily_hours,
        "totals":             totals,
        "totals_match":       totals_match,
        "warnings":           warnings,
        "expenses_cad":       expenses_cad,
        "expenses_usd":       expenses_usd,
        "expense_totals_cad": totals_cad,
        "expense_totals_usd": totals_usd,
    }
