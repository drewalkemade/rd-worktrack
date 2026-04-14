"""
Timesheet Extractor
Extracts employee timesheet data from Excel files (.xlsx format)
"""

import openpyxl
from datetime import datetime
from typing import List, Dict, Optional


class TimesheetExtractor:
    """Extractor for Excel timesheet files"""

    def __init__(self):
        """Initialize the extractor"""
        self.sheet_name = 'Biweekly Time Sheet'
        self.name_cell = 'H3'
        self.end_date_cell = 'H5'
        self.totals_row = 23

        # Column mappings for row 23 totals
        self.column_mapping = {
            'D': 'Regular Hours',
            'E': 'Overtime 1 Hours',
            'F': 'Overtime 2 Hours',
            'G': 'Drive Hours',
            'H': 'Sick/PEL',
            'I': 'Vacation',
            'J': 'Holiday',
            'K': 'Non-Billable'
        }

    def extract_timesheet(self, file_path: str) -> Optional[Dict]:
        """
        Extract data from a single timesheet

        Args:
            file_path: Path to the Excel timesheet file

        Returns:
            Dictionary containing employee timesheet data or None if error
        """
        try:
            # Load workbook with data_only=True to get values instead of formulas
            wb = openpyxl.load_workbook(file_path, data_only=True)

            if self.sheet_name not in wb.sheetnames:
                return {
                    'error': f"Sheet '{self.sheet_name}' not found in {file_path}",
                    'file': file_path
                }

            ws = wb[self.sheet_name]

            # Extract employee name
            name = ws[self.name_cell].value
            if not name:
                return {
                    'error': f"No name found in cell {self.name_cell}",
                    'file': file_path
                }

            # Extract end date
            end_date = ws[self.end_date_cell].value
            if not end_date:
                return {
                    'error': f"No end date found in cell {self.end_date_cell}",
                    'file': file_path
                }

            # Convert end_date to datetime if it's not already
            if isinstance(end_date, str):
                try:
                    end_date = datetime.strptime(end_date, '%Y-%m-%d')
                except ValueError:
                    try:
                        end_date = datetime.strptime(end_date, '%m/%d/%Y')
                    except ValueError:
                        return {
                            'error': f"Invalid date format in {self.end_date_cell}: {end_date}",
                            'file': file_path
                        }

            # Extract hours from row 23
            hours_data = {}
            for col_letter, label in self.column_mapping.items():
                cell_ref = f'{col_letter}{self.totals_row}'
                value = ws[cell_ref].value

                # Convert to float, default to 0 if None or invalid
                try:
                    hours_data[label] = float(value) if value is not None else 0.0
                except (ValueError, TypeError):
                    hours_data[label] = 0.0

            return {
                'name': name.strip(),
                'end_date': end_date,
                'hours': hours_data,
                'file': file_path
            }

        except Exception as e:
            return {
                'error': f"Error processing {file_path}: {str(e)}",
                'file': file_path
            }

    def extract_multiple_timesheets(self, file_paths: List[str]) -> Dict:
        """
        Extract data from multiple timesheets

        Args:
            file_paths: List of paths to Excel timesheet files

        Returns:
            Dictionary with 'data' (list of successful extractions) and 'errors' (list of errors)
        """
        results = {
            'data': [],
            'errors': []
        }

        for file_path in file_paths:
            result = self.extract_timesheet(file_path)

            if result and 'error' in result:
                results['errors'].append(result)
            elif result:
                results['data'].append(result)

        return results

    def sort_by_name_order(self, data: List[Dict], name_order: List[str]) -> List[Dict]:
        """
        Sort data by specific name order

        Args:
            data: List of employee timesheet data
            name_order: Ordered list of names to sort by

        Returns:
            Sorted list of timesheet data
        """
        # Create a mapping of name to sort order
        order_map = {name: i for i, name in enumerate(name_order)}

        # Sort by the order, putting unmatched names at the end
        def sort_key(item):
            name = item['name']
            return order_map.get(name, len(name_order))

        return sorted(data, key=sort_key)


# Testing function
if __name__ == "__main__":
    extractor = TimesheetExtractor()

    # Test with sample timesheets
    import glob

    timesheet_files = sorted(glob.glob("*.xlsx"))
    timesheet_files = [f for f in timesheet_files if not f.startswith('test_export')]

    print(f"Found {len(timesheet_files)} timesheet files")
    print("=" * 80)

    results = extractor.extract_multiple_timesheets(timesheet_files)

    print(f"\nSuccessfully extracted: {len(results['data'])} timesheets")
    print(f"Errors: {len(results['errors'])}")

    if results['data']:
        print("\n--- Extracted Data ---")
        for entry in results['data']:
            print(f"\nName: {entry['name']}")
            print(f"End Date: {entry['end_date'].strftime('%Y-%m-%d')}")
            print(f"Hours: {entry['hours']}")

    if results['errors']:
        print("\n--- Errors ---")
        for error in results['errors']:
            print(f"\nFile: {error['file']}")
            print(f"Error: {error['error']}")
