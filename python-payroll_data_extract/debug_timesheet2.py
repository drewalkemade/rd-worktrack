"""
Debug script to examine timesheet structure with actual values
"""

import openpyxl
from datetime import datetime

def examine_timesheet(filename):
    print(f"Examining: {filename}")
    print("=" * 80)

    try:
        wb = openpyxl.load_workbook(filename, data_only=True)  # data_only=True gets values not formulas

        if 'Biweekly Time Sheet' in wb.sheetnames:
            ws = wb['Biweekly Time Sheet']

            print("\n--- Key Cells (Values) ---")
            print(f"H3 (Name): {ws['H3'].value}")
            print(f"H4 (Start Date?): {ws['H4'].value}")
            print(f"H5 (End Date): {ws['H5'].value}")

            print("\n--- Headers Row 8 ---")
            for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                cell = ws[f'{col}8']
                print(f"{col}8: {cell.value}")

            print("\n--- Row 23 (Total hours - Values) ---")
            for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                cell = ws[f'{col}23']
                print(f"{col}23: {cell.value}")

        else:
            print("\nSheet 'Biweekly Time Sheet' not found!")

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # Test with all timesheets
    timesheets = [
        "DTrifTS_17_20251123.xlsx",
        "FMoldovan_20251123.xlsx",
        "JJeremiasTS_16_20251123.xlsx",
        "PaulRobertson_44.11232025.xlsx",
        "Saleh_Y_2025-11-23.xlsx",
        "ZacharyEbbinghaus_20251123.xlsx"
    ]

    for ts in timesheets:
        examine_timesheet(ts)
        print("\n")
