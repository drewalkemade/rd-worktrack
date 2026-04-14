"""
Debug script to examine timesheet structure
"""

import openpyxl
import sys

def examine_timesheet(filename):
    print(f"Examining: {filename}")
    print("=" * 80)

    try:
        wb = openpyxl.load_workbook(filename)

        print(f"\nSheet names: {wb.sheetnames}")

        # Look for 'Biweekly Time Sheet' sheet
        if 'Biweekly Time Sheet' in wb.sheetnames:
            ws = wb['Biweekly Time Sheet']

            print("\n--- Key Cells ---")
            print(f"H3 (Name): {ws['H3'].value}")
            print(f"H5 (Date): {ws['H5'].value}")

            print("\n--- Row 23 ---")
            row_23_values = []
            for col in range(1, 20):  # Check first 20 columns
                cell = ws.cell(row=23, column=col)
                if cell.value is not None:
                    col_letter = openpyxl.utils.get_column_letter(col)
                    row_23_values.append(f"{col_letter}23: {cell.value}")

            for val in row_23_values:
                print(val)

            print("\n--- Row 22 (headers?) ---")
            row_22_values = []
            for col in range(1, 20):
                cell = ws.cell(row=22, column=col)
                if cell.value is not None:
                    col_letter = openpyxl.utils.get_column_letter(col)
                    row_22_values.append(f"{col_letter}22: {cell.value}")

            for val in row_22_values:
                print(val)

        else:
            print("\nSheet 'Biweekly Time Sheet' not found!")
            print("Available sheets:", wb.sheetnames)

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # Test with first timesheet
    timesheets = [
        "DTrifTS_17_20251123.xlsx",
        "FMoldovan_20251123.xlsx",
        "JJeremiasTS_16_20251123.xlsx"
    ]

    for ts in timesheets[:1]:  # Just examine first one
        examine_timesheet(ts)
